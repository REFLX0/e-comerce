import openpyxl
import json

wb = openpyxl.load_workbook(r'oil\Tomobile_Inventory_Complete.xlsx')
all_products = []

for sheet_name in ['Mannol', 'Liqui Moly', 'Osram', 'Neolux']:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[4]]
    print(f'{sheet_name}: headers = {headers}')

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[1]:
            continue
        prod = {
            'title': row[1],
            'sku': row[2] or '',
            'brand': row[3] or sheet_name,
            'price_tnd': str(row[4]) if row[4] else '0',
            'stock': row[5] or '',
            'categories': row[6] or '',
            'short_description': row[7] or '',
            'image_url': row[8] or '',
            'url': row[9] or '',
            'full_description': row[10] or '',
        }
        all_products.append(prod)

print(f'\nTotal products from Excel: {len(all_products)}')
with_price = [p for p in all_products if p['price_tnd'] and p['price_tnd'] != '0']
print(f'Products with price: {len(with_price)}')

for p in with_price[:5]:
    t = p['title']
    pr = p['price_tnd']
    print(f'  {t} -> {pr}')

with open(r'oil\excel_products.json', 'w', encoding='utf-8') as f:
    json.dump(all_products, f, ensure_ascii=False, indent=2)
print('Saved to oil/excel_products.json')
